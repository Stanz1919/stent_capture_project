"""
data.comsol_loader
==================
Loader for COMSOL gradient-decay datasets cleaned by the user and exported to
``data/comsol_gradient_data_cleaned.xlsx`` at the project root.

**Important label correction.** The sheets whose titles read ``1 T (normalised)``
were actually run at **B₀ = 1.5 T** in COMSOL; the sheet titled ``0.15 T`` was
actually run at **B₀ = 0.2433 T**. Labels were not corrected in the source
spreadsheet — this loader applies the correction in the ``applied_B_true``
attribute on every :class:`ComsolDataset`.

Each sheet is a cut-line export from a 3-D FEM model of the V-series stent
geometries (6, 10, 12, 18 cells per loop), taken from a strut edge radially
inward. The gradient magnitude is reported in T/m as a function of distance
from the strut surface (in mm). Trendlines were fitted in Excel to the form
``G(d) = A × d^n``.

Public API
----------

- :func:`load_dataset(key)` — return a :class:`ComsolDataset` by name.
- :func:`all_datasets()` — yield every dataset in canonical order.
- :data:`SHEETS` — ordered mapping of dataset key → sheet name.

Use from figure modules::

    from stent_capture.data.comsol_loader import load_dataset
    v2 = load_dataset("V2")
    d_mm, grad = v2.d_mm, v2.grad_T_per_m
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterator

import numpy as np

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise ImportError(
        "openpyxl is required to read the COMSOL dataset. "
        "Install with: pip install openpyxl"
    ) from exc


_DATA_XLSX = Path(__file__).parents[2] / "data" / "comsol_gradient_data_cleaned.xlsx"

SHEETS: dict[str, str] = {
    "V1":      "V1_6cell_cleaned",
    "V2":      "V2_12cell_cleaned",
    "V3":      "V3_18cell_cleaned",
    "V4_new":  "V5_10cell_NEW_cleaned",
    "2D_1T":   "2D_1T_cleaned",
    "2D_015T": "2D_015T_cleaned",
}

N_STRUTS: dict[str, int | None] = {
    "V1": 6, "V2": 12, "V3": 18, "V4_new": 10, "2D_1T": None, "2D_015T": None,
}

APPLIED_B_TRUE: dict[str, float] = {
    "V1":      1.5,
    "V2":      1.5,
    "V3":      1.5,
    "V4_new":  1.5,
    "2D_1T":   1.5,
    "2D_015T": 0.2433,
}


@dataclass
class ComsolDataset:
    """Cleaned COMSOL gradient-decay dataset (single cut line)."""

    key: str
    sheet: str
    n_struts: int | None
    applied_B_label: str
    applied_B_true: float                    # T — label-corrected
    d_mm: np.ndarray                         # distance from strut surface (mm)
    grad_T_per_m: np.ndarray
    fit_A: float | None = None
    fit_n: float | None = None
    fit_R2: float | None = None
    fit_range_mm: tuple[float, float] | None = None
    notes: str = ""

    @property
    def d_m(self) -> np.ndarray:
        """Distance from strut surface in metres."""
        return self.d_mm * 1e-3

    def fit_predict(self, d_mm: np.ndarray) -> np.ndarray:
        """Evaluate the Excel-fitted power law at *d_mm*."""
        if self.fit_A is None or self.fit_n is None:
            raise ValueError(f"{self.key}: no fit available")
        return self.fit_A * np.asarray(d_mm, dtype=float) ** self.fit_n


def _first_num_col_indices(header_row: tuple) -> dict[str, int]:
    """Map canonical column names to their 0-based index within the sheet."""
    mapping: dict[str, int] = {}
    for i, cell in enumerate(header_row):
        if not isinstance(cell, str):
            continue
        h = cell.strip().lower()
        if h.startswith("d from strut"):
            mapping["d"] = i
        elif h.startswith("|") and "b" in h:
            mapping["grad"] = i
    return mapping


def _parse_sheet(wb, key: str) -> ComsolDataset:
    sheet_name = SHEETS[key]
    ws = wb[sheet_name]

    header_idx: int | None = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if any(isinstance(v, str) and v.strip().lower().startswith("d from strut")
               for v in row):
            header_idx = i
            header = row
            break
    if header_idx is None:
        raise ValueError(f"Could not locate header row in sheet {sheet_name!r}")

    col = _first_num_col_indices(header)
    if "d" not in col or "grad" not in col:
        raise ValueError(
            f"Sheet {sheet_name!r}: missing 'd from strut' or '|∇B|' column"
        )

    d_vals: list[float] = []
    g_vals: list[float] = []
    for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
        d_raw = row[col["d"]] if col["d"] < len(row) else None
        g_raw = row[col["grad"]] if col["grad"] < len(row) else None
        if d_raw is None or g_raw is None:
            continue
        try:
            d_vals.append(float(d_raw))
            g_vals.append(float(g_raw))
        except (TypeError, ValueError):
            continue

    d_mm = np.asarray(d_vals, dtype=float)
    grad = np.asarray(g_vals, dtype=float)

    order = np.argsort(d_mm)
    d_mm = d_mm[order]
    grad = grad[order]

    return ComsolDataset(
        key=key,
        sheet=sheet_name,
        n_struts=N_STRUTS[key],
        applied_B_label="1 T (norm.)" if key != "2D_015T" else "0.15 T",
        applied_B_true=APPLIED_B_TRUE[key],
        d_mm=d_mm,
        grad_T_per_m=grad,
    )


def _extract_fit_from_summary(wb) -> dict[str, dict]:
    ws = wb["Summary"]
    fits: dict[str, dict] = {}

    header_idx: int | None = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row and isinstance(row[0], str) and row[0].strip() == "Dataset":
            header_idx = i
            break
    if header_idx is None:
        return fits

    key_by_row = {
        "V1": "V1", "V2": "V2", "V5": "V4_old", "V5 (new)": "V4_new",
        "V3": "V3", "2D": "2D_1T",
    }
    seen_2d = False
    for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
        if not row or row[0] is None:
            continue
        label = str(row[0]).strip()
        if label not in key_by_row:
            continue
        key = key_by_row[label]
        if label == "2D":
            if seen_2d:
                continue
            seen_2d = True
        try:
            A = float(row[4])
            n_raw = row[5]
            if isinstance(n_raw, str):
                n = float(n_raw.split()[0])
            else:
                n = float(n_raw)
            R2 = float(row[8]) if row[8] is not None else None
        except (TypeError, ValueError, IndexError):
            continue
        fits[key] = {"A": A, "n": n, "R2": R2}
    return fits

_CACHE: dict[str, ComsolDataset] = {}


def load_dataset(key: str) -> ComsolDataset:
    """Load (and cache) a single dataset by key (see :data:`SHEETS`)."""
    if key in _CACHE:
        return _CACHE[key]
    if key not in SHEETS:
        raise KeyError(f"Unknown dataset {key!r}; valid: {list(SHEETS)}")
    if not _DATA_XLSX.exists():
        raise FileNotFoundError(f"COMSOL dataset not found at {_DATA_XLSX}")

    wb = load_workbook(_DATA_XLSX, data_only=True, read_only=True)
    try:
        ds = _parse_sheet(wb, key)
        fits = _extract_fit_from_summary(wb)
    finally:
        wb.close()

    if key == "V4_new":
        ds.fit_range_mm = (0.0, 0.17)
        ds.notes = "near-field fit only (d < 0.17 mm)"
    f = fits.get(key)
    if f is not None:
        ds.fit_A = f["A"]
        ds.fit_n = f["n"]
        ds.fit_R2 = f["R2"]

    _CACHE[key] = ds
    return ds


def all_datasets() -> Iterator[ComsolDataset]:
    """Yield every dataset in canonical order."""
    for key in SHEETS:
        yield load_dataset(key)


def summary_table() -> list[dict]:
    """Return a list of summary dicts — useful for quick CLI inspection."""
    out = []
    for ds in all_datasets():
        out.append(dict(
            key=ds.key, n_struts=ds.n_struts, B_true=ds.applied_B_true,
            n_points=len(ds.d_mm),
            A=ds.fit_A, n=ds.fit_n, R2=ds.fit_R2,
        ))
    return out


if __name__ == "__main__":
    for row in summary_table():
        print(row)
